import time
import struct
import can
from openpyxl import load_workbook

def send_float64_over_can(bus, message_id, value):
    # Convierte el valor float a 64-bit (double precision) en formato bytes
    packed_value = struct.pack('>d', value)  # '>d' para big-endian double
    
    # Crea el mensaje CAN con un identificador extendido (29 bits)
    message = can.Message(
        arbitration_id=message_id,
        data=packed_value,
        is_extended_id=True  # Usar identificadores extendidos (29 bits) para CAN 2.0B
    )
    
    # Intentar enviar el mensaje hasta que sea exitoso
    while True:
        try:
            bus.send(message)
            break  # Si se envía correctamente, salir del bucle
        except can.CanOperationError as e:
            print(f"Error al enviar el mensaje CAN con ID {message_id}: {e}")
            time.sleep(0.1)  # Espera antes de reintentar

def read_all_columns(file_path, sheet_name):
    # Carga el archivo XLSX
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Lee todos los valores y los organiza por filas
    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Excluye la primera fila si son encabezados
    
    # Transforma las filas en columnas
    columns = list(zip(*rows))
    return columns

def get_can_id(system_prefix, sensor_index):
    # Calcula el CAN ID basado en el prefijo del sistema y el índice del sensor
    # Para CAN 2.0B, el identificador puede ser más largo (hasta 29 bits)
    return (system_prefix << 24) | (sensor_index & 0xFFFFFF)

def main():
    # Configura la interfaz CAN
    bus = can.interface.Bus(channel='can0', bustype='socketcan')  # Ajusta según tu configuración

    # Parámetros del archivo y hoja
    file_path = '/home/yazminbc/Downloads/EVsimdata.xlsx'  # Ruta al archivo cargado
    sheet_name = 'Hoja1'  # Nombre de la hoja, ajusta si es necesario

    # Definición de los sistemas y sus rangos de columnas
    systems = {
        0x1: range(0, 5),   # Battery System: columnas A-E (0-4)
        0x2: range(5, 9),   # Motor System: columnas F-I (5-8)
        0x3: range(9, 12),  # Driveline System: columnas J-L (9-11)
        0x4: range(12, 14)  # Glider System: columnas M-N (12-13)
    }

    # Lee todos los datos por columna
    columns = read_all_columns(file_path, sheet_name)

    # Intervalo de tiempo entre envíos en segundos (0.5 segundos para 2Hz)
    interval = 0.5
    
    num_rows = len(columns[0])  # Número de filas (asume que todas las columnas tienen el mismo número de filas)
    
    for row_index in range(num_rows):
        start_time = time.monotonic()  # Marca el tiempo de inicio del ciclo
        
        for system_prefix, column_range in systems.items():
            for sensor_index in column_range:
                value = columns[sensor_index][row_index]
                if value is not None:
                    can_id = get_can_id(system_prefix, sensor_index - column_range.start)
                    send_float64_over_can(bus, can_id, value)
        
        # Calcular el tiempo que falta para completar el intervalo de 0.5 segundos
        elapsed_time = time.monotonic() - start_time
        time_to_wait = interval - elapsed_time
        
        if time_to_wait > 0:
            time.sleep(time_to_wait)  # Espera el tiempo restante si es necesario

if __name__ == "__main__":
    main()

