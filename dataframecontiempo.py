import time
import struct
import can
from openpyxl import load_workbook
import threading

# Lock para sincronizar el envío de mensajes
lock = threading.Lock()

def send_float32_with_timestamp_over_can(bus, message_id, timestamp, value, max_retries=5, retry_delay=0.1):
    # Convertir el timestamp y el valor en float32 (4 bytes cada uno)
    timestamp_float32 = struct.pack('>f', timestamp)
    value_float32 = struct.pack('>f', value)
    
    # Empaquetar ambos valores (8 bytes en total: 4 bytes para el timestamp y 4 bytes para el valor)
    packed_data = timestamp_float32 + value_float32

    message = can.Message(
        arbitration_id=message_id,
        data=packed_data,
        is_extended_id=True
    )
    
    retries = 0
    while retries < max_retries:
        try:
            # Enviar el mensaje con el timestamp y el valor
            bus.send(message)
            return True
        except can.CanOperationError as e:
            print(f"Error al enviar el mensaje CAN con ID {message_id}: {e}")
            retries += 1
            time.sleep(retry_delay)
    
    return False

def read_all_columns(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    columns = list(zip(*rows))
    return columns

def get_can_id(system_prefix, sensor_index):
    return (system_prefix << 24) | (sensor_index & 0xFFFFFF)

def send_can_message_with_timestamp(bus, can_id, value):
    timestamp = time.time()  # Obtener el timestamp actual
    timestamp_float32 = float(timestamp)  # Convertir el timestamp a float32
    value_float32 = float(value)  # Asegurarse de que el valor sea float32
    
    with lock:  # Asegura que solo un thread envíe un mensaje a la vez
        success = send_float32_with_timestamp_over_can(bus, can_id, timestamp_float32, value_float32)
        if not success:
            print(f"Failed to send message with CAN ID {can_id}")

def process_row(bus, columns, row_index, systems):
    for system_prefix, column_range in systems.items():
        for sensor_index in column_range:
            value = columns[sensor_index][row_index]
            if value is not None:
                can_id = get_can_id(system_prefix, sensor_index - column_range.start)
                thread = threading.Thread(target=send_can_message_with_timestamp, args=(bus, can_id, value))
                thread.start()
                thread.join()  # Espera hasta que el mensaje actual sea enviado antes de continuar

def main():
    bus = can.interface.Bus(channel='can0', bustype='socketcan')
    file_path = '/home/yazminbc/Downloads/EVsimdata.xlsx'
    sheet_name = 'Hoja1'

    systems = {
        0x1: range(0, 5),   # Battery System: columnas A-E (0-4)
        0x2: range(5, 9),   # Motor System: columnas F-I (5-8)
        0x3: range(9, 12),  # Driveline System: columnas J-L (9-11)
        0x4: range(12, 13)  # Glider System: columna M
    }

    columns = read_all_columns(file_path, sheet_name)
    interval = 0.2  # 5Hz = 1/5 segundos = 200ms
    num_rows = len(columns[0])

    # Factor para seleccionar filas (aproximadamente cada 3 filas)
    row_selection_factor = 3

    for row_index in range(0, num_rows, row_selection_factor):
        start_time = time.monotonic()
        
        # Procesa la fila seleccionada
        process_row(bus, columns, row_index, systems)
        
        # Calcula el tiempo restante para completar los 200ms
        elapsed_time = time.monotonic() - start_time
        time_to_wait = interval - elapsed_time
        
        if time_to_wait > 0:
            time.sleep(time_to_wait)

if __name__ == "__main__":
    main()
